import datetime
import os
import sys
from openpyxl import load_workbook, Workbook

# ==========================================
# LER MODELO
# ==========================================
def carregar_modelo(caminho):
    wb = load_workbook(caminho)
    ws = wb.active

    atividades = []

    for row in ws.iter_rows(min_row=10, values_only=True):
        tipo = row[0]
        ordem = row[1]
        data = row[2]
        hora = row[4]
        nome = row[5]

        if not tipo or not ordem or not data:
            continue

        if hora is None:
            hora = datetime.time(0, 0)

        if isinstance(hora, str):
            h, m = map(int, hora.split(":"))
            hora = datetime.time(h, m)

        dt = datetime.datetime.combine(data, hora)

        atividades.append({
            "tipo": tipo,
            "ordem": int(ordem),
            "nome": nome,
            "datetime": dt
        })

    return atividades


# ==========================================
# AGRUPAR
# ==========================================
def agrupar_por_ordem(atividades):
    grupos = {}
    for a in atividades:
        grupos.setdefault(a["ordem"], []).append(a)
    return dict(sorted(grupos.items()))


# ==========================================
# CALCULAR OFFSETS
# ==========================================
def calcular_offsets(grupos):
    base_ordem = min(grupos.keys())
    base_data = grupos[base_ordem][0]["datetime"]

    offsets = {}

    for ordem, lista in grupos.items():
        diff = (lista[0]["datetime"] - base_data).days
        offsets[ordem] = diff

    return offsets


# ==========================================
# ALERTAS CALENDÁRIO
# ==========================================
def verificar_alertas(dt, dias_sem_expediente):
    alertas = []
    dia_semana = dt.weekday()

    if dia_semana in [5, 6]:
        alertas.append("⚠️ Final de semana")

    if (dt.day, dt.month) in dias_sem_expediente:
        alertas.append("⚠️ Dia sem expediente")

    if dia_semana == 4 and not (8 <= dt.hour < 12):
        alertas.append("⚠️ Sexta fora do expediente")

    if alertas:
        print("\n".join(alertas))
        while True:
            op = input("Deseja prosseguir mesmo assim? (S/N): ").strip().upper()
            if op in ["S", "N"]:
                return op == "S"

    return True


# ==========================================
# INPUT DATA
# ==========================================
def ler_data_hora(ano):
    while True:
        try:
            data = input("Data (dd/mm): ")
            dia, mes = map(int, data.split("/"))

            hora = input("Hora (HH:MM): ")
            h, mi = map(int, hora.split(":"))

            return datetime.datetime(ano, mes, dia, h, mi)
        except:
            print("❌ Data inválida.")


# ==========================================
# LER DATAS P
# ==========================================
def ler_datas_principais(grupos, ano, dias_sem_expediente):
    datas_p = {}
    lista_p = []
    ultima_data = None

    for ordem, lista in grupos.items():
        for item in lista:
            if item["tipo"] == "P":
                lista_p.append((ordem, item["nome"]))

    i = 0

    print("\n=== INFORME AS DATAS DAS ATIVIDADES PRINCIPAIS ===\n")

    while i < len(lista_p):
        ordem, nome = lista_p[i]
        print(f"{ordem} - {nome}")

        dt = ler_data_hora(ano)

        if not verificar_alertas(dt, dias_sem_expediente):
            continue

        if ultima_data and dt < ultima_data:
            print("⚠️ Data anterior à anterior.")
            continue

        datas_p[ordem] = dt
        ultima_data = dt
        i += 1

    return datas_p


# ==========================================
# GERAR CRONOGRAMA (AJUSTADO)
# ==========================================
def gerar_cronograma(grupos, offsets, datas_p):
    novo = []

    ordem_base = list(datas_p.keys())[0]
    data_base = datas_p[ordem_base]

    ultima_data_gerada = None

    for ordem, lista in grupos.items():

        # 🔹 Se tem principal → usa exatamente a data dela
        if ordem in datas_p:
            data_referencia = datas_p[ordem]

        else:
            # 🔹 Mantém lógica de offset
            deslocamento = offsets[ordem] - offsets[ordem_base]
            data_referencia = data_base + datetime.timedelta(days=deslocamento)

        for item in lista:
            nova_data = data_referencia.replace(
                hour=item["datetime"].hour,
                minute=item["datetime"].minute
            )

            # 🔹 Garante ordem cronológica (não volta no tempo)
            if ultima_data_gerada and nova_data < ultima_data_gerada:
                nova_data = ultima_data_gerada

            nova = {
                "ordem": ordem,
                "tipo": item["tipo"],
                "nome": item["nome"],
                "datetime": nova_data
            }

            novo.append(nova)
            ultima_data_gerada = nova_data

    return sorted(novo, key=lambda x: (x["ordem"], x["datetime"]))


# ==========================================
# EXPORTAR EXCEL
# ==========================================
def exportar_excel(atividades, caminho_modelo):
    wb = load_workbook(caminho_modelo)
    ws = wb.active

    i = 0

    for row in ws.iter_rows(min_row=10):
        if i >= len(atividades):
            break

        a = atividades[i]

        def get_celula_principal(cell):
            for merged in ws.merged_cells.ranges:
                if cell.coordinate in merged:
                    return ws.cell(merged.min_row, merged.min_col)
            return cell

        cell_data = get_celula_principal(row[2])
        cell_hora = get_celula_principal(row[4])

        cell_data.value = a["datetime"].date()
        cell_hora.value = a["datetime"].time()

        i += 1

    if getattr(sys, 'frozen', False):
        pasta_execucao = os.path.dirname(sys.executable)
    else:
        pasta_execucao = os.path.dirname(os.path.abspath(__file__))

    caminho_saida = os.path.join(pasta_execucao, "cronograma_gerado.xlsx")

    wb.save(caminho_saida)
    os.startfile(caminho_saida)


# ==========================================
# CALENDÁRIO
# ==========================================
def preparar_calendario():
    dias = set()
    while True:
        entrada = input("Data sem expediente (dd/mm ou ENTER): ")
        if not entrada:
            break
        try:
            d, m = map(int, entrada.split("/"))
            dias.add((d, m))
        except:
            print("Inválido")
    return dias


# ==========================================
# MAIN
# ==========================================
def main():
    print("=== CRONOGRAMA AUTOMÁTICO ===\n")

    ano = int(input("Ano: "))
    dias_sem_expediente = preparar_calendario()

    pasta = os.path.dirname(__file__)
    caminho = os.path.join(pasta, "Modelo cronograma.xlsx")

    atividades = carregar_modelo(caminho)
    grupos = agrupar_por_ordem(atividades)
    offsets = calcular_offsets(grupos)

    datas_p = ler_datas_principais(grupos, ano, dias_sem_expediente)

    novo = gerar_cronograma(grupos, offsets, datas_p)
    exportar_excel(novo, caminho)

    print("\n✔ Concluído!")


if __name__ == "__main__":
    main()